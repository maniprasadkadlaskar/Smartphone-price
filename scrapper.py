import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
ws.title = "Smartphone price"
ws.append( ["Smartphone" , "Price in INR"] )
ws.cell(row = 1 , column = 1).font = Font(size = 12 , name = "Arial")
ws.cell(row = 1 , column = 2).font = Font(size = 12 , name = "Arial")
ws.column_dimensions['A'].width = 70
ws.column_dimensions['B'].width = 30
url = "https://www.flipkart.com/search?q=smartphone&as=on&as-show=on&otracker=AS_Query_OrganicAutoSuggest_5_5_na_na_na&otracker1=AS_Query_OrganicAutoSuggest_5_5_na_na_na&as-pos=5&as-type=RECENT&suggestionId=smartphone&requestId=7412ef77-d6e8-49c1-9579-f5687e094dcf&as-backfill=on"

req = requests.get(url)

soup = BeautifulSoup(req.text , "html.parser")

price = soup.find_all(class_="_30jeq3 _1_WHN1")
mobileName = soup.find_all(class_ ="_4rR01T")
length  = len(price)

i = 0
while i < length:
    ws.append( [mobileName[i].string , price[i].string] )
    i += 1

wb.save("data.xlsx")